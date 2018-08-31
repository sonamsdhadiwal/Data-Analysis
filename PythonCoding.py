import openpyxl
import xlrd

#open excel document with above module
wb=openpyxl.load_workbook('Password.xlsx')
sheet1=wb.get_sheet_by_name('Sheet1')
N = int(input('Enter the number of samples for testing'))
verficationSamples = 400-N
print('Number of samples for verification are'+str(verficationSamples))
sum1H_period = 0
sum1DD_period_t=0
sum1UD_period_t = 0

sum1H_t =0
sum1DD_t_i =0
sum1UD_t_i =0

sum1H_i =0
sum1DD_i_e =0
sum1UD_i_e =0

sum1H_e =0
sum1DD_e_five=0
sum1UD_e_five =0

sum1H_five =0
sum1DD_five_Shift_r =0
sum1UD_five_Shift_r=0

sum1H_Shift_r =0
sum1DD_Shift_r_o =0
sum1UD_Shift_r_o =0

sum1H_o =0
sum1DD_o_a =0
sum1UD_o_a =0

sum1H_0 =0
sum1DD_a_n =0
sum1UD_a_n =0

sum1H_n =0
sum1DD_n_l =0
sum1UD_n_l =0

sum1H_l =0
sum1DD_l_Return =0
sum1UD_l_Return =0
sum1H_Return =0
count=0
rangeOfN = N*50
#calculating mean of each time period
for i in range(2,rangeOfN,50):
    count=count+1
    sum1H_period = sum1H_period+sheet1.cell(row=i,column=4).value
    sum1DD_period_t = sum1DD_period_t+sheet1.cell(row=i,column=5).value
    sum1UD_period_t = sum1UD_period_t+sheet1.cell(row=i,column=6).value
    
    sum1H_t = sum1H_t+sheet1.cell(row=i,column=7).value
    sum1DD_t_i = sum1DD_t_i+sheet1.cell(row=i,column=8).value
    sum1UD_t_i = sum1UD_t_i+sheet1.cell(row=i,column=9).value
    
    sum1H_i = sum1H_i+sheet1.cell(row=i,column=10).value
    sum1DD_i_e = sum1DD_i_e+sheet1.cell(row=i,column=11).value
    sum1UD_i_e = sum1UD_i_e+sheet1.cell(row=i,column=12).value
    
    sum1H_e = sum1H_e+sheet1.cell(row=i,column=13).value
    sum1DD_e_five = sum1DD_e_five+sheet1.cell(row=i,column=14).value
    sum1UD_e_five = sum1UD_e_five+sheet1.cell(row=i,column=15).value
    
    sum1H_five = sum1H_five+sheet1.cell(row=i,column=16).value
    sum1DD_five_Shift_r = sum1DD_five_Shift_r+sheet1.cell(row=i,column=17).value
    sum1UD_five_Shift_r = sum1UD_five_Shift_r+sheet1.cell(row=i,column=18).value
    
    sum1H_Shift_r = sum1H_Shift_r+sheet1.cell(row=i,column=19).value
    sum1DD_Shift_r_o = sum1DD_Shift_r_o+sheet1.cell(row=i,column=20).value
    sum1UD_Shift_r_o = sum1UD_Shift_r_o+sheet1.cell(row=i,column=21).value

    sum1H_o = sum1H_o+sheet1.cell(row=i,column=22).value
    sum1DD_o_a = sum1DD_o_a+sheet1.cell(row=i,column=23).value
    ssum1UD_o_a = sum1UD_Shift_r_o+sheet1.cell(row=i,column=24).value

    sum1H_0 = sum1H_0+sheet1.cell(row=i,column=25).value
    sum1DD_a_n = sum1DD_a_n+sheet1.cell(row=i,column=26).value
    sum1UD_a_n = sum1UD_a_n+sheet1.cell(row=i,column=27).value
    
    sum1H_n = sum1H_n+sheet1.cell(row=i,column=28).value
    sum1DD_n_l = sum1DD_n_l+sheet1.cell(row=i,column=29).value
    sum1UD_n_l = sum1UD_n_l+sheet1.cell(row=i,column=30).value

    sum1H_l = sum1H_l+sheet1.cell(row=i,column=31).value
    sum1DD_l_Return = sum1DD_l_Return+sheet1.cell(row=i,column=32).value
    sum1UD_l_Return = sum1UD_l_Return+sheet1.cell(row=i,column=33).value
    sum1H_Return = sum1H_Return+sheet1.cell(row=i,column=34).value
    
Avg_sum1H_period = sum1H_period/count
Avg_sum1DD_period_t = (sum1DD_period_t/count)
Avg_sum1UD_period_t = (sum1UD_period_t/count)

Avg_sum1H_t =(sum1H_t/count)
Avg_sum1DD_t_i = (sum1DD_t_i/count)
Avg_sum1UD_t_i =(sum1UD_t_i/count)

Avg_sum1H_i =(sum1H_i/count)
Avg_sum1DD_i_e =(sum1DD_i_e/count)
Avg_sum1UD_i_e = (sum1UD_i_e/count)

Avg_sum1H_e =(sum1H_e/count)
Avg_sum1DD_e_five =(sum1DD_e_five/count)
Avg_sum1UD_e_five = (sum1UD_e_five/count)

Avg_sum1H_five = (sum1H_five/count)
Avg_sum1DD_five_Shift_r =(sum1DD_five_Shift_r/count)
Avg_sum1UD_five_Shift_r = (sum1UD_five_Shift_r/count)

Avg_sum1H_Shift_r = (sum1H_Shift_r/count)
Avg_sum1DD_Shift_r_o =(sum1DD_Shift_r_o/count)
Avg_sum1UD_Shift_r_o =(sum1UD_Shift_r_o/count)

Avg_sum1H_o = (sum1H_o/count)
Avg_sum1DD_o_a =(sum1DD_o_a/count)
Avg_sum1UD_o_a =(sum1UD_o_a/count)

Avg_sum1H_0 = (sum1H_0/count)
Avg_sum1DD_a_n =(sum1DD_a_n/count)
Avg_sum1UD_a_n = (sum1UD_a_n/count)

Avg_sum1H_n =(sum1H_n/count)
Avg_sum1DD_n_l =(sum1DD_n_l/count)
Avg_sum1UD_n_l = (sum1UD_n_l/count)

Avg_sum1H_l = (sum1H_l/count)
Avg_sum1DD_l_Return =(sum1DD_l_Return/count)
Avg_sum1UD_l_Return =(sum1UD_l_Return/count)
Avg_sum1H_Return = (sum1H_Return/count)


#getting probe vectors
#for user1
N = (rangeOfN)+2
temp=(rangeOfN)+2
print('Printing genuine scores')
countUser=0
GenuineScoresUser1=[]
#while (countUser!=52 and N!=20401):          
for i in range(N,20401,50):
    countUser=countUser+1
    GenuineScores = (abs(Avg_sum1H_period-sheet1.cell(row=i,column=4).value) +
                 abs(Avg_sum1DD_period_t-sheet1.cell(row=i,column=5).value)+
                 abs(Avg_sum1UD_period_t-sheet1.cell(row=i,column=6).value)+
                 abs(Avg_sum1H_t -sheet1.cell(row=i,column=7).value) +
                 abs(Avg_sum1DD_t_i - sheet1.cell(row=i,column=8).value)+
                 abs(Avg_sum1UD_t_i-sheet1.cell(row=i,column=9).value)+
                 abs(Avg_sum1H_i-sheet1.cell(row=i,column=10).value)+
                 abs(Avg_sum1DD_i_e - sheet1.cell(row=i,column=11).value)+
                 abs(Avg_sum1UD_i_e - sheet1.cell(row=i,column=12).value)+
                 abs(Avg_sum1H_e - sheet1.cell(row=i,column=13).value)+
                 abs(Avg_sum1DD_e_five - sheet1.cell(row=i,column=14).value)+
                 abs(Avg_sum1UD_e_five - sheet1.cell(row=i,column=15).value)+
                 abs(Avg_sum1H_five - sheet1.cell(row=i,column=16).value)+
                 abs(Avg_sum1DD_five_Shift_r - sheet1.cell(row=i,column=17).value)+
                 abs(Avg_sum1UD_five_Shift_r - sheet1.cell(row=i,column=18).value)+
                 abs(Avg_sum1H_Shift_r-sheet1.cell(row=i,column=19).value)+
                 abs(Avg_sum1DD_Shift_r_o-sheet1.cell(row=i,column=20).value)+
                 abs(Avg_sum1UD_Shift_r_o-sheet1.cell(row=i,column=21).value)+
                 abs(Avg_sum1H_o-sheet1.cell(row=i,column=22).value)+
                 abs(Avg_sum1DD_o_a-sheet1.cell(row=i,column=23).value)+
                 abs(Avg_sum1UD_o_a-sheet1.cell(row=i,column=24).value)+
                 abs(Avg_sum1H_0-sheet1.cell(row=i,column=25).value)+
                 abs(Avg_sum1DD_a_n-sheet1.cell(row=i,column=26).value)+
                 abs(Avg_sum1UD_a_n-sheet1.cell(row=i,column=27).value)+
                 abs(Avg_sum1H_n-sheet1.cell(row=i,column=28).value)+
                 abs(Avg_sum1DD_n_l-sheet1.cell(row=i,column=29).value)+
                 abs(Avg_sum1UD_n_l-sheet1.cell(row=i,column=30).value)+
                 abs(Avg_sum1H_l - sheet1.cell(row=i,column=31).value)+
                 abs(Avg_sum1DD_l_Return - sheet1.cell(row=i,column=32).value)+
                 abs(Avg_sum1UD_l_Return-sheet1.cell(row=i,column=33).value)+
                 abs(Avg_sum1H_Return-sheet1.cell(row=i,column=34).value))
    GenuineScores = round(GenuineScores/31,2)
    GenuineScoresUser1.append(GenuineScores)
    N=N+1
print('Genuine Scores are')
print(GenuineScoresUser1)

count1=0
count2=0
threshold = float(input('Enter the threshold value'))
length = len(GenuineScoresUser1)
for i in range(0,length):
    if(GenuineScoresUser1[i]<threshold):
        count1=count1+1
    else:
        count2 = count2 +1
#print count1
#print count2
falseRejectRate =float(count2)/float(length)
print(falseRejectRate)

#Calculating impostors
ImpostorScoresUser1=[]
count3=0
count4=0
N=temp+1
#while (N!=20400):
for i in range(N,20400):
    if(i%50)==N:
        count3=count3+1
    else:
        count4=count4+1
        ImpostorScores= (abs(Avg_sum1H_period-sheet1.cell(row=i,column=4).value) +
                 abs(Avg_sum1DD_period_t-sheet1.cell(row=i,column=5).value)+
                 abs(Avg_sum1UD_period_t-sheet1.cell(row=i,column=6).value)+
                 abs(Avg_sum1H_t -sheet1.cell(row=i,column=7).value) +
                 abs(Avg_sum1DD_t_i - sheet1.cell(row=i,column=8).value)+
                 abs(Avg_sum1UD_t_i-sheet1.cell(row=i,column=9).value)+
                 abs(Avg_sum1H_i-sheet1.cell(row=i,column=10).value)+
                 abs(Avg_sum1DD_i_e - sheet1.cell(row=i,column=11).value)+
                 abs(Avg_sum1UD_i_e - sheet1.cell(row=i,column=12).value)+
                 abs(Avg_sum1H_e - sheet1.cell(row=i,column=13).value)+
                 abs(Avg_sum1DD_e_five - sheet1.cell(row=i,column=14).value)+
                 abs(Avg_sum1UD_e_five - sheet1.cell(row=i,column=15).value)+
                 abs(Avg_sum1H_five - sheet1.cell(row=i,column=16).value)+
                 abs(Avg_sum1DD_five_Shift_r - sheet1.cell(row=i,column=17).value)+
                 abs(Avg_sum1UD_five_Shift_r - sheet1.cell(row=i,column=18).value)+
                 abs(Avg_sum1H_Shift_r-sheet1.cell(row=i,column=19).value)+
                 abs(Avg_sum1DD_Shift_r_o-sheet1.cell(row=i,column=20).value)+
                 abs(Avg_sum1UD_Shift_r_o-sheet1.cell(row=i,column=21).value)+
                 abs(Avg_sum1H_o-sheet1.cell(row=i,column=22).value)+
                 abs(Avg_sum1DD_o_a-sheet1.cell(row=i,column=23).value)+
                 abs(Avg_sum1UD_o_a-sheet1.cell(row=i,column=24).value)+
                 abs(Avg_sum1H_0-sheet1.cell(row=i,column=25).value)+
                 abs(Avg_sum1DD_a_n-sheet1.cell(row=i,column=26).value)+
                 abs(Avg_sum1UD_a_n-sheet1.cell(row=i,column=27).value)+
                 abs(Avg_sum1H_n-sheet1.cell(row=i,column=28).value)+
                 abs(Avg_sum1DD_n_l-sheet1.cell(row=i,column=29).value)+
                 abs(Avg_sum1UD_n_l-sheet1.cell(row=i,column=30).value)+
                 abs(Avg_sum1H_l - sheet1.cell(row=i,column=31).value)+
                 abs(Avg_sum1DD_l_Return - sheet1.cell(row=i,column=32).value)+
                 abs(Avg_sum1UD_l_Return-sheet1.cell(row=i,column=33).value)+
                 abs(Avg_sum1H_Return-sheet1.cell(row=i,column=34).value))
        ImpostorScores = round(ImpostorScores/31,2)
        ImpostorScoresUser1.append(ImpostorScores)
N=N+1
print('Impostor Scores are')
print(ImpostorScoresUser1)

count5=0
count6=0
length1= len(ImpostorScoresUser1)
for i in range(0,length1):
    if(ImpostorScoresUser1[i]<threshold):
        count5=count5+1
    else:
        count6 = count6+1
#print count5
#print count6
falseAcceptRate = float(count6)/length1
print('False Accept rate is')
print(falseAcceptRate)

