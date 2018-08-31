import openpyxl
import xlrd

#open excel document with above module
wb=openpyxl.load_workbook('Password.xlsx')
sheet1=wb.get_sheet_by_name('Sheet1')
N = int(input('Enter the number of samples for testing'))
verficationSamples = 400-N
print('Number of samples for verification are'+str(verficationSamples))

    
#Calculate mean hold time for each user
'''col1=[]
col2=[]
for row in range(1,sheet1.max_row+1):
    col1.append((sheet1.cell(row,4).value)              
    print(col1)
#print(col2)
'''
'''
for row in range(2,5):
    sheet1['AI1']= 'SUM(D1:AH1)'
    sheet1['AI2']= '=SUM(D2:AH2)'
    wb.save('Password.xlsx')
'''
'''
workbook = xlrd.open_workbook('Password.xlsx')
worksheet = workbook.sheet_by_name('Sheet1')
num_rows = worksheet.nrows - 1
curr_row = 0

#creates an array to store all the rows
row_array = []

while curr_row < num_rows:
    row = worksheet.row(curr_row)
    row_array += row
    curr_row += 1

print row_array
'''

#hold period
hDot=[0,0]
hT=[0,0]
hI=[0,0]
hE=[0,0]
hFive=[0,0]
hR=[0,0]
hO=[0,0]
hA=[0,0]
hN=[0,0]
h1=[0,0]
hAverage=[0,0]
for i in range(2,20401):
     hDot.append(sheet1.cell(row=i, column=4).value)
     hT.append(sheet1.cell(row=i, column=7).value)
     hI.append(sheet1.cell(row=i, column=10).value)
     hE.append(sheet1.cell(row=i, column=13).value)
     hFive.append(sheet1.cell(row=i, column=16).value)
     hR.append(sheet1.cell(row=i, column=19).value)
     hO.append(sheet1.cell(row=i, column=22).value)
     hA.append(sheet1.cell(row=i, column=25).value)
     hN.append(sheet1.cell(row=i, column=28).value)
     h1.append(sheet1.cell(row=i, column=31).value)

#DD period
ddDot=[0,0]
ddT=[0,0]
ddI=[0,0]
ddE=[0,0]
ddFive=[0,0]
ddR=[0,0]
ddO=[0,0]
ddA=[0,0]
ddN=[0,0]
dd1=[0,0]
ddAverage=[0,0]
for i in range(2,20401):
     ddDot.append(sheet1.cell(row=i, column=5).value)
     ddT.append(sheet1.cell(row=i, column=8).value)
     ddI.append(sheet1.cell(row=i, column=11).value)
     ddE.append(sheet1.cell(row=i, column=14).value)
     ddFive.append(sheet1.cell(row=i, column=17).value)
     ddR.append(sheet1.cell(row=i, column=20).value)
     ddO.append(sheet1.cell(row=i, column=23).value)
     ddA.append(sheet1.cell(row=i, column=26).value)
     ddN.append(sheet1.cell(row=i, column=29).value)
     dd1.append(sheet1.cell(row=i, column=32).value)


#UD time
udDot=[0,0]
udT=[0,0]
udI=[0,0]
udE=[0,0]
udFive=[0,0]
udR=[0,0]
udO=[0,0]
udA=[0,0]
udN=[0,0]
ud1=[0,0]
udAverage=[0,0]
for i in range(2,20401):
     udDot.append(sheet1.cell(row=i, column=6).value)
     udT.append(sheet1.cell(row=i, column=9).value)
     udI.append(sheet1.cell(row=i, column=12).value)
     udE.append(sheet1.cell(row=i, column=15).value)
     udFive.append(sheet1.cell(row=i, column=18).value)
     udR.append(sheet1.cell(row=i, column=21).value)
     udO.append(sheet1.cell(row=i, column=24).value)
     udA.append(sheet1.cell(row=i, column=27).value)
     udN.append(sheet1.cell(row=i, column=30).value)
     ud1.append(sheet1.cell(row=i, column=33).value)

for j in range(2,20401):
     hAverage = (hDot[j]+hT[j]+hI[j]+hE[j]+hFive[j]+hR[j]+hO[j]+hA[j]+hN[j]+h1[j])/10
     ddAverage = (ddDot[j]+ddT[j]+ddI[j]+ddE[j]+ddFive[j]+ddR[j]+ddO[j]+ddA[j]+ddN[j]+dd1[j])/10
     udAverage = (udDot[j]+udT[j]+udI[j]+udE[j]+udFive[j]+udR[j]+udO[j]+udA[j]+udN[j]+ud1[j])/10
    # print (hAverage,ddAverage,udAverage)
     #Calculating trials of users
     trial = 0.0
     trial1 = 0.0
     trial2 = 0.0
     count =0
     #this will give number trials of each user
     #Dividing by total number of samples will give mean of users
#Calculate only for number of N typing samples
     l=2
     while(l<N):
         for k in range(l,20401,400):
             count = count+1
             trial =float(trial)+float(hAverage)
             trial1 = float(trial1)+float(ddAverage)
             trial2 = float(trial2) + float(udAverage)
     #will be getting total 51 trial--mean template of each user-- total 51 mean templates
     # total 400 templates
         print ((trial)/51,trial1/51,trial2/52)
     l = l+1
     
#TODO: Get mean of these trials-taken above

#calculating probes
trial3 = 0.0
trial4 = 0.0
trial5 = 0.0
totalLength=N+1
     while(totalLength<401):
         for k in range(totalLength,20401,400):
             count1 = count1+1
             trial3 =float(trial3)+float(hAverage)
             trial4 = float(trial4)+float(ddAverage)
             trial5 = float(trial5) + float(udAverage)
     #will be getting total 51 trial--mean template of each user-- total 51 mean templates
     # total 400 templates
         print ((trial3)/51,trial4/51,trial5/51)
     totalLength = totalLength+1

#TODO:calculate genuine score using manhattan distance formula

#write everything in excel sheet which will make calculation easier

#TODO: Calculate zero impostor scores for each user using manhattan distance formula
     
#TODO: Calculate FAR and FRR
